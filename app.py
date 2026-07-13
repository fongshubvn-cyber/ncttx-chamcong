import os
import io
import sqlite3
import datetime
from flask import Flask, request, jsonify, render_template, send_file, url_for
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import get_db, init_db

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join('static', 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload size

# Đảm bảo các thư mục tồn tại
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Khởi động database khi app chạy
with app.app_context():
    init_db()

# --- Helper Functions ---
def get_user_by_code(code):
    conn = get_db()
    user = conn.execute('''
        SELECT u.*, a.name as area_name 
        FROM users u 
        LEFT JOIN areas a ON u.area_id = a.id 
        WHERE u.code = ? AND u.status = 'active'
    ''', (code,)).fetchone()
    conn.close()
    return user

def calculate_time_duration(check_in_str, check_out_str):
    # Trả về số giờ làm việc giữa 2 mốc thời gian dạng string ISO
    try:
        fmt = "%Y-%m-%d %H:%M:%S"
        t1 = datetime.datetime.strptime(check_in_str, fmt)
        t2 = datetime.datetime.strptime(check_out_str, fmt)
        diff = t2 - t1
        return diff.total_seconds() / 3600.0
    except Exception:
        return 0.0

def get_staff_kpi_data(start_date=None, end_date=None, search_query=None):
    conn = get_db()
    
    # Lấy tất cả nhân sự (trừ manager)
    if search_query:
        q = f"%{search_query}%"
        users = conn.execute("""
            SELECT u.*, a.name as area_name 
            FROM users u 
            LEFT JOIN areas a ON u.area_id = a.id 
            WHERE u.role != 'manager' AND (u.name LIKE ? OR u.code LIKE ?)
        """, (q, q)).fetchall()
    else:
        users = conn.execute("SELECT u.*, a.name as area_name FROM users u LEFT JOIN areas a ON u.area_id = a.id WHERE u.role != 'manager'").fetchall()
    
    # Lấy time logs
    if start_date and end_date:
        time_query = "SELECT * FROM time_logs WHERE date(timestamp) >= ? AND date(timestamp) <= ? ORDER BY timestamp ASC"
        time_logs = conn.execute(time_query, (start_date, end_date)).fetchall()
    else:
        time_query = "SELECT * FROM time_logs ORDER BY timestamp ASC"
        time_logs = conn.execute(time_query).fetchall()
        
    # Lấy submission checklist
    if start_date and end_date:
        sub_query = "SELECT s.*, u.name as grader_name, a.name as area_name FROM checklist_submissions s JOIN users u ON s.grader_id = u.id JOIN areas a ON s.area_id = a.id WHERE date(s.timestamp) >= ? AND date(s.timestamp) <= ?"
        submissions = conn.execute(sub_query, (start_date, end_date)).fetchall()
    else:
        sub_query = "SELECT s.*, u.name as grader_name, a.name as area_name FROM checklist_submissions s JOIN users u ON s.grader_id = u.id JOIN areas a ON s.area_id = a.id"
        submissions = conn.execute(sub_query).fetchall()
        
    # Lấy chi tiết các checklist
    if start_date and end_date:
        details_query = "SELECT d.*, s.area_id, s.status as sub_status FROM checklist_submission_details d JOIN checklist_submissions s ON d.submission_id = s.id WHERE date(s.timestamp) >= ? AND date(s.timestamp) <= ?"
        submission_details = conn.execute(details_query, (start_date, end_date)).fetchall()
    else:
        details_query = "SELECT d.*, s.area_id, s.status as sub_status FROM checklist_submission_details d JOIN checklist_submissions s ON d.submission_id = s.id"
        submission_details = conn.execute(details_query).fetchall()
    
    conn.close()
    
    # Tính toán thời gian làm việc
    user_hours = {}
    user_status = {}  # Đang check-in hay out
    
    for u in users:
        user_hours[u['id']] = 0.0
        user_status[u['id']] = 'checked_out'
    
    # Ghép cặp check-in / check-out của từng user
    last_check_in = {}
    for log in time_logs:
        uid = log['user_id']
        action = log['action']
        ts = log['timestamp']
        
        if action == 'check_in':
            last_check_in[uid] = ts
            user_status[uid] = 'checked_in'
        elif action == 'check_out' and uid in last_check_in:
            duration = calculate_time_duration(last_check_in[uid], ts)
            user_hours[uid] += duration
            del last_check_in[uid]
            user_status[uid] = 'checked_out'
 
    # Nếu hiện tại vẫn đang check-in, tính tạm thời gian đến hiện tại (chỉ nếu hôm nay nằm trong khoảng lọc)
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    today_str = datetime.datetime.now().strftime("%Y-%m-%d")
    for uid, ts in last_check_in.items():
        if start_date and end_date:
            if not (start_date <= today_str <= end_date):
                continue
        duration = calculate_time_duration(ts, now_str)
        user_hours[uid] += duration

    # Tính toán KPIs vệ sinh khu vực phụ trách
    # Đối với mỗi user, khu vực phụ trách là u.area_id.
    # Ta xem các checklist chấm cho khu vực đó có bao nhiêu mục Đạt (pass) và Không đạt (fail).
    # Chỉ số được quyết định bởi đánh giá của Quản lý:
    # - Nếu Quản lý từ chối (rejected): toàn bộ hạng mục bị coi là KHÔNG đạt (fail)
    # - Nếu Quản lý duyệt đạt (approved) hoặc chưa duyệt (pending): tính theo đánh giá của nhân viên đi chấm
    area_pass_fail = {}
    for d in submission_details:
        aid = d['area_id']
        sub_status = d['sub_status']
        item_status = d['status']
        
        is_pass = False
        if sub_status == 'rejected':
            # Bị từ chối làm lại: tất cả tính là fail
            is_pass = False
        else:
            # pending hoặc approved: tính theo điểm chấm
            is_pass = (item_status == 'pass')
            
        if aid not in area_pass_fail:
            area_pass_fail[aid] = {'pass': 0, 'total': 0}
        area_pass_fail[aid]['total'] += 1
        if is_pass:
            area_pass_fail[aid]['pass'] += 1

    # Tính tỷ lệ chấm chéo đã nộp (Checklist Completion)
    # Nhân viên phải chấm chéo các khu vực khác. 
    # Giả định mỗi ca làm việc (mỗi lượt check-in), nhân sự cần thực hiện 1 checklist chấm chéo.
    # Tỷ lệ hoàn thành = (số checklist đã chấm) / (số lần check-in).
    user_checkins = {}
    for log in time_logs:
        if log['action'] == 'check_in':
            user_checkins[log['user_id']] = user_checkins.get(log['user_id'], 0) + 1

    user_submissions = {}
    for sub in submissions:
        gid = sub['grader_id']
        user_submissions[gid] = user_submissions.get(gid, 0) + 1

    report_list = []
    for u in users:
        uid = u['id']
        aid = u['area_id']
        
        # 1. Điểm vệ sinh của khu vực do user phụ trách
        hygiene_score = 100.0
        total_graded_items = 0
        if aid and aid in area_pass_fail:
            total_graded_items = area_pass_fail[aid]['total']
            if total_graded_items > 0:
                hygiene_score = (area_pass_fail[aid]['pass'] / total_graded_items) * 100.0
        
        # 2. Tỷ lệ hoàn thành nhiệm vụ chấm chéo
        checkins_count = user_checkins.get(uid, 0)
        submitted_count = user_submissions.get(uid, 0)
        
        if checkins_count > 0:
            completion_rate = min((submitted_count / checkins_count) * 100.0, 100.0)
        else:
            # Nếu chưa check-in bao giờ nhưng đã chấm (chạy thử), cho 100%
            completion_rate = 100.0 if submitted_count > 0 else 0.0

        # Phân loại hiệu suất
        # Hiệu quả: Điểm vệ sinh khu vực >= 80% VÀ tỷ lệ chấm chéo >= 80% (hoặc không yêu cầu nếu chưa check-in)
        # Không hiệu quả: Điểm vệ sinh < 80% HOẶC tỷ lệ chấm chéo < 60% (nếu đã từng check-in)
        is_efficient = True
        efficiency_reason = []
        
        if aid:
            if total_graded_items > 0 and hygiene_score < 80.0:
                is_efficient = False
                efficiency_reason.append(f"Khu vực phụ trách bẩn ({hygiene_score:.1f}% đạt)")
        
        if checkins_count > 0 and completion_rate < 70.0:
            is_efficient = False
            efficiency_reason.append(f"Ít hoàn thành chấm chéo ({completion_rate:.1f}%)")

        status_text = "Hiệu quả" if is_efficient else "Cần cải thiện"
        reason_text = ", ".join(efficiency_reason) if efficiency_reason else "Hoạt động tốt"

        report_list.append({
            'user_id': uid,
            'name': u['name'],
            'code': u['code'],
            'role': u['role'],
            'area_name': u['area_name'] or "Chưa phân bổ",
            'total_hours': round(user_hours[uid], 2),
            'checkins_count': checkins_count,
            'checklists_submitted': submitted_count,
            'hygiene_score': round(hygiene_score, 1),
            'completion_rate': round(completion_rate, 1),
            'status': status_text,
            'reason': reason_text,
            'current_status': user_status[uid]
        })
        
    return report_list

# --- Routes ---

@app.route('/')
def index():
    return render_template('index.html')

# 1. API: Authentication
@app.route('/api/login', methods=['POST'])
def login():
    data = request.json or {}
    code = data.get('code', '').strip().upper()
    if not code:
        return jsonify({'error': 'Vui lòng nhập mã đăng nhập'}), 400
    
    user = get_user_by_code(code)
    if not user:
        return jsonify({'error': 'Mã đăng nhập không đúng hoặc tài khoản bị khóa'}), 401
    
    return jsonify({
        'id': user['id'],
        'name': user['name'],
        'code': user['code'],
        'role': user['role'],
        'area_id': user['area_id'],
        'area_name': user['area_name'] or ''
    })

# 2. API: Time Logs (Chấm công)
@app.route('/api/time_logs', methods=['GET'])
def get_user_time_logs():
    user_id = request.args.get('user_id', type=int)
    if not user_id:
        return jsonify({'error': 'Thiếu user_id'}), 400
        
    conn = get_db()
    logs = conn.execute('''
        SELECT l.*, a.name as area_name
        FROM time_logs l
        LEFT JOIN areas a ON l.area_id = a.id
        WHERE l.user_id = ? AND date(l.timestamp) = date('now', 'localtime')
        ORDER BY l.timestamp DESC
    ''', (user_id,)).fetchall()
    conn.close()
    return jsonify([dict(log) for log in logs])

@app.route('/api/time_logs/checkout_preview', methods=['GET'])
def time_log_checkout_preview():
    user_id = request.args.get('user_id', type=int)
    if not user_id:
        return jsonify({'error': 'Thiếu user_id'}), 400
        
    conn = get_db()
    
    # Thời gian hiện tại format SQLite
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # 1. Tìm check-in gần nhất trong ngày để tính thời gian làm việc
    last_checkin = conn.execute('''
        SELECT timestamp FROM time_logs 
        WHERE user_id = ? AND action = 'check_in' AND date(timestamp) = date('now', 'localtime')
        ORDER BY timestamp DESC LIMIT 1
    ''', (user_id,)).fetchone()
    
    duration_str = "Chưa xác định"
    if last_checkin:
        duration_hours = calculate_time_duration(last_checkin['timestamp'], now_str)
        total_minutes = int(duration_hours * 60)
        h = total_minutes // 60
        m = total_minutes % 60
        if h > 0:
            duration_str = f"{h} giờ {m} phút"
        else:
            duration_str = f"{m} phút"
            
    # 2. Tìm checklist ca của ngày hôm nay
    submission = conn.execute('''
        SELECT s.id, a.name as area_name FROM checklist_submissions s
        JOIN areas a ON s.area_id = a.id
        WHERE s.grader_id = ? AND date(s.timestamp) = date('now', 'localtime')
        ORDER BY s.timestamp DESC LIMIT 1
    ''', (user_id,)).fetchone()
    
    checklist_summary = None
    if submission:
        details = conn.execute('''
            SELECT status FROM checklist_submission_details 
            WHERE submission_id = ?
        ''', (submission['id'],)).fetchall()
        
        total_items = len(details)
        passed_items = sum(1 for d in details if d['status'] == 'pass')
        percent = (passed_items / total_items) * 100 if total_items > 0 else 0
        
        checklist_summary = {
            'done': True,
            'area_name': submission['area_name'],
            'total': total_items,
            'passed': passed_items,
            'percent': round(percent, 1)
        }
    else:
        checklist_summary = {
            'done': False,
            'percent': 0.0,
            'total': 0,
            'passed': 0,
            'area_name': 'Chưa thực hiện'
        }
        
    conn.close()
    
    return jsonify({
        'working_time': duration_str,
        'checklist': checklist_summary
    })


@app.route('/api/time_logs/status', methods=['GET'])
def time_log_status():
    user_id = request.args.get('user_id', type=int)
    if not user_id:
        return jsonify({'error': 'Thiếu user_id'}), 400
    
    conn = get_db()
    # Tìm log chấm công cuối cùng của user trong ngày hôm nay
    last_log = conn.execute('''
        SELECT * FROM time_logs 
        WHERE user_id = ? AND date(timestamp) = date('now', 'localtime')
        ORDER BY timestamp DESC LIMIT 1
    ''', (user_id,)).fetchone()
    
    # Tìm thông tin khu vực mặc định của user để tra cứu trạng thái vệ sinh
    user = conn.execute("SELECT area_id FROM users WHERE id = ?", (user_id,)).fetchone()
    
    area_hygiene = {'status': 'none', 'notes': '', 'grader': ''}
    if user and user['area_id']:
        area_sub = conn.execute('''
            SELECT s.status, s.manager_notes, u.name as grader_name 
            FROM checklist_submissions s
            JOIN users u ON s.grader_id = u.id
            WHERE s.area_id = ? AND date(s.timestamp) = date('now', 'localtime')
            ORDER BY s.timestamp DESC LIMIT 1
        ''', (user['area_id'],)).fetchone()
        if area_sub:
            area_hygiene = {
                'status': area_sub['status'],
                'notes': area_sub['manager_notes'] or '',
                'grader': area_sub['grader_name']
            }
            
    conn.close()
    
    response_data = {}
    if last_log:
        response_data = {
            'status': last_log['action'],
            'timestamp': last_log['timestamp'],
            'area_id': last_log['area_id'],
            'area_hygiene': area_hygiene
        }
    else:
        response_data = {
            'status': 'checked_out',
            'timestamp': None,
            'area_id': None,
            'area_hygiene': area_hygiene
        }
        
    return jsonify(response_data)

@app.route('/api/time_logs/toggle', methods=['POST'])
def time_log_toggle():
    data = request.json or {}
    user_id = data.get('user_id')
    action = data.get('action') # 'check_in' hoặc 'check_out'
    area_id = data.get('area_id') # Gửi kèm khi check_in
    latitude = data.get('latitude')
    longitude = data.get('longitude')
    
    if not user_id or not action:
        return jsonify({'error': 'Thiếu tham số'}), 400
        
    conn = get_db()
    
    # Kiểm tra user có tồn tại không
    user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        conn.close()
        return jsonify({'error': 'Nhân viên không tồn tại'}), 404
        
    # Thời gian hiện tại format SQLite
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    if action == 'check_in':
        if not area_id:
            conn.close()
            return jsonify({'error': 'Vui lòng chọn khu vực làm việc ca này'}), 400
        # Ghi nhận check-in kèm toạ độ GPS
        conn.execute('INSERT INTO time_logs (user_id, action, timestamp, area_id, latitude, longitude) VALUES (?, ?, ?, ?, ?, ?)',
                     (user_id, 'check_in', now_str, area_id, latitude, longitude))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'action': action, 'timestamp': now_str})
    else:
        # 1. Tìm check-in gần nhất trong ngày để tính thời gian làm việc
        last_checkin = conn.execute('''
            SELECT timestamp FROM time_logs 
            WHERE user_id = ? AND action = 'check_in' AND date(timestamp) = date('now', 'localtime')
            ORDER BY timestamp DESC LIMIT 1
        ''', (user_id,)).fetchone()
        
        duration_str = "Chưa xác định"
        if last_checkin:
            duration_hours = calculate_time_duration(last_checkin['timestamp'], now_str)
            total_minutes = int(duration_hours * 60)
            h = total_minutes // 60
            m = total_minutes % 60
            if h > 0:
                duration_str = f"{h} giờ {m} phút"
            else:
                duration_str = f"{m} phút"
        
        # 2. Tìm checklist ca của ngày hôm nay
        submission = conn.execute('''
            SELECT s.id, a.name as area_name FROM checklist_submissions s
            JOIN areas a ON s.area_id = a.id
            WHERE s.grader_id = ? AND date(s.timestamp) = date('now', 'localtime')
            ORDER BY s.timestamp DESC LIMIT 1
        ''', (user_id,)).fetchone()
        
        checklist_summary = None
        if submission:
            details = conn.execute('''
                SELECT status FROM checklist_submission_details 
                WHERE submission_id = ?
            ''', (submission['id'],)).fetchall()
            
            total_items = len(details)
            passed_items = sum(1 for d in details if d['status'] == 'pass')
            percent = (passed_items / total_items) * 100 if total_items > 0 else 0
            
            checklist_summary = {
                'done': True,
                'area_name': submission['area_name'],
                'total': total_items,
                'passed': passed_items,
                'percent': round(percent, 1)
            }
        else:
            checklist_summary = {
                'done': False,
                'percent': 0.0,
                'total': 0,
                'passed': 0,
                'area_name': 'Chưa thực hiện'
            }
        # Ghi nhận check-out kèm toạ độ GPS
        conn.execute('INSERT INTO time_logs (user_id, action, timestamp, area_id, latitude, longitude) VALUES (?, ?, ?, NULL, ?, ?)',
                     (user_id, 'check_out', now_str, latitude, longitude))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True, 
            'action': action, 
            'timestamp': now_str,
            'checkout_summary': {
                'working_time': duration_str,
                'checklist': checklist_summary
            }
        })

# 3. API: Upload ảnh
@app.route('/api/uploads/image', methods=['POST'])
def upload_image():
    if 'image' not in request.files:
        return jsonify({'error': 'Không tìm thấy file ảnh'}), 400
    
    file = request.files['image']
    if file.filename == '':
        return jsonify({'error': 'Tên file trống'}), 400
        
    filename = secure_filename(file.filename)
    # Thêm timestamp vào tên file tránh trùng lặp
    ext = os.path.splitext(filename)[1]
    filename = f"img_{int(datetime.datetime.now().timestamp())}{ext}"
    
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(file_path)
    
    # Trả về đường dẫn tĩnh tương đối
    url = f"/static/uploads/{filename}"
    return jsonify({'url': url})

# 4. API: Checklist Chấm chéo
@app.route('/api/checklist/assigned', methods=['GET'])
def checklist_assigned():
    grader_id = request.args.get('grader_id', type=int)
    if not grader_id:
        return jsonify({'error': 'Thiếu grader_id'}), 400
        
    conn = get_db()
    # Lấy thông tin user
    user = conn.execute("SELECT * FROM users WHERE id = ?", (grader_id,)).fetchone()
    if not user:
        conn.close()
        return jsonify({'error': 'Nhân viên không tồn tại'}), 404
        
    # Tìm khu vực làm việc hiện tại của nhân sự (trong ca check-in hôm nay)
    current_log = conn.execute('''
        SELECT area_id FROM time_logs 
        WHERE user_id = ? AND action = 'check_in' AND date(timestamp) = date('now', 'localtime')
        ORDER BY timestamp DESC LIMIT 1
    ''', (grader_id,)).fetchone()
    
    user_current_area = current_log['area_id'] if current_log else user['area_id']
    
    # Lấy danh sách tất cả các khu vực khác khu vực hiện tại của nhân sự (chấm chéo)
    other_areas = conn.execute("SELECT * FROM areas WHERE id != ?", (user_current_area,)).fetchall()
    
    # Nếu không có khu vực khác (chỉ có duy nhất 1 khu vực), đành phải chấm chính khu vực đó
    if not other_areas:
        other_areas = conn.execute("SELECT * FROM areas").fetchall()
        
    if not other_areas:
        conn.close()
        return jsonify({'error': 'Chưa cấu hình khu vực nào trên hệ thống'}), 404
        
    # Chọn khu vực để chấm chéo: 
    # 1. Kiểm tra xem có quy tắc phân công chấm chéo nào cho khu vực hiện tại của user hay không
    rule = conn.execute("SELECT to_area_id FROM cross_check_rules WHERE from_area_id = ?", (user_current_area,)).fetchone()
    
    selected_area = None
    if rule:
        # Nếu có quy tắc, lấy thông tin khu vực mục tiêu
        target_area_id = rule['to_area_id']
        selected_area = conn.execute("SELECT * FROM areas WHERE id = ?", (target_area_id,)).fetchone()
        
    # 2. Nếu không có quy tắc, fallback về thuật toán chọn khu vực có số lượt chấm ít nhất trong ngày
    if not selected_area:
        min_submissions = 999999
        for area in other_areas:
            count = conn.execute('''
                SELECT COUNT(*) FROM checklist_submissions 
                WHERE area_id = ? AND date(timestamp) = date('now', 'localtime')
            ''', (area['id'],)).fetchone()[0]
            
            if count < min_submissions:
                min_submissions = count
                selected_area = area
            
    # Lấy danh sách checklist của khu vực được chọn
    items = conn.execute("SELECT * FROM checklist_items WHERE area_id = ?", (selected_area['id'],)).fetchall()
    
    # Kiểm tra xem nhân sự này hôm nay đã chấm khu vực này chưa
    existing_sub = conn.execute('''
        SELECT * FROM checklist_submissions 
        WHERE grader_id = ? AND area_id = ? AND date(timestamp) = date('now', 'localtime')
        ORDER BY timestamp DESC LIMIT 1
    ''', (grader_id, selected_area['id'])).fetchone()
    
    submission_data = None
    if existing_sub:
        details = conn.execute('''
            SELECT d.*, i.task_name, i.reference_image
            FROM checklist_submission_details d
            JOIN checklist_items i ON d.item_id = i.id
            WHERE d.submission_id = ?
        ''', (existing_sub['id'],)).fetchall()
        
        submission_data = {
            'id': existing_sub['id'],
            'status': existing_sub['status'],
            'manager_notes': existing_sub['manager_notes'],
            'details': [dict(d) for d in details]
        }
        
    conn.close()
    
    return jsonify({
        'area': {
            'id': selected_area['id'],
            'name': selected_area['name'],
            'description': selected_area['description']
        },
        'items': [dict(item) for item in items],
        'existing_submission': submission_data
    })

@app.route('/api/checklist/submit_complete', methods=['POST'])
def checklist_submit_complete():
    data = request.json or {}
    grader_id = data.get('grader_id')
    area_id = data.get('area_id')
    items = data.get('items', []) # List of {item_id, status, captured_image, notes}
    latitude = data.get('latitude')
    longitude = data.get('longitude')
    
    if not grader_id or not area_id or not items:
        return jsonify({'error': 'Dữ liệu chấm không hợp lệ'}), 400
        
    conn = get_db()
    
    try:
        # Tạo bản ghi submission mới (hoặc ghi đè lượt chấm trong ngày của người này cho khu vực này)
        # Xóa lượt chấm cũ trong ngày để ghi nhận lượt chấm mới nhất
        existing = conn.execute('''
            SELECT id FROM checklist_submissions 
            WHERE grader_id = ? AND area_id = ? AND date(timestamp) = date('now', 'localtime')
        ''', (grader_id, area_id)).fetchone()
        
        if existing:
            conn.execute("DELETE FROM checklist_submissions WHERE id = ?", (existing['id'],))
            
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO checklist_submissions (grader_id, area_id, timestamp, status, latitude, longitude)
            VALUES (?, ?, ?, 'pending', ?, ?)
        ''', (grader_id, area_id, now_str, latitude, longitude))
        
        sub_id = cursor.lastrowid
        
        # Thêm chi tiết
        for item in items:
            conn.execute('''
                INSERT INTO checklist_submission_details (submission_id, item_id, status, captured_image, notes)
                VALUES (?, ?, ?, ?, ?)
            ''', (sub_id, item['item_id'], item['status'], item['captured_image'], item.get('notes', '')))
            
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'submission_id': sub_id})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': f'Lỗi hệ thống: {str(e)}'}), 500

# 5. API: Manager Dashboard & Reports
@app.route('/api/manager/dashboard', methods=['GET'])
def manager_dashboard():
    conn = get_db()
    # Nhân sự đang làm việc (check-in cuối cùng trong ngày chưa check-out)
    # Lấy check-in mới nhất của mỗi user trong ngày, kiểm tra xem có phải là check_in không
    active_staff_count = 0
    users = conn.execute("SELECT id, name FROM users WHERE role != 'manager'").fetchall()
    
    for u in users:
        last_log = conn.execute('''
            SELECT action FROM time_logs 
            WHERE user_id = ? AND date(timestamp) = date('now', 'localtime')
            ORDER BY timestamp DESC LIMIT 1
        ''', (u['id'],)).fetchone()
        if last_log and last_log['action'] == 'check_in':
            active_staff_count += 1
            
    # Tống số lượt chấm checklist trong ngày hôm nay
    today_checklist_count = conn.execute('''
        SELECT COUNT(*) FROM checklist_submissions 
        WHERE date(timestamp) = date('now', 'localtime')
    ''').fetchone()[0]
    
    # Tổng số khu vực và nhân viên
    total_areas = conn.execute("SELECT COUNT(*) FROM areas").fetchone()[0]
    total_staff = len(users)
    
    # Lấy toàn bộ logs để ghép ca làm việc (Shifts)
    logs = conn.execute('''
        SELECT l.*, u.name as user_name, a.name as area_name
        FROM time_logs l
        JOIN users u ON l.user_id = u.id
        LEFT JOIN areas a ON l.area_id = a.id
        ORDER BY l.user_id, l.timestamp ASC
    ''').fetchall()
    
    shifts_by_user = {}
    for log in logs:
        uid = log['user_id']
        action = log['action']
        ts = log['timestamp']
        area = log['area_name']
        uname = log['user_name']
        
        if uid not in shifts_by_user:
            shifts_by_user[uid] = []
            
        if action == 'check_in':
            # Bắt đầu ca mới
            shifts_by_user[uid].append({
                'user_name': uname,
                'area_name': area or '-',
                'check_in': ts,
                'check_in_lat': log['latitude'],
                'check_in_lng': log['longitude'],
                'check_out': None,
                'check_out_lat': None,
                'check_out_lng': None,
                'duration': None
            })
        elif action == 'check_out':
            # Ghép vào ca check_in gần nhất chưa check_out
            if shifts_by_user[uid] and shifts_by_user[uid][-1]['check_out'] is None:
                shifts_by_user[uid][-1]['check_out'] = ts
                shifts_by_user[uid][-1]['check_out_lat'] = log['latitude']
                shifts_by_user[uid][-1]['check_out_lng'] = log['longitude']
                duration = calculate_time_duration(shifts_by_user[uid][-1]['check_in'], ts)
                shifts_by_user[uid][-1]['duration'] = f"{duration:.2f} giờ"
            else:
                # Trường hợp không có check_in tương ứng
                shifts_by_user[uid].append({
                    'user_name': uname,
                    'area_name': area or '-',
                    'check_in': None,
                    'check_in_lat': None,
                    'check_in_lng': None,
                    'check_out': ts,
                    'check_out_lat': log['latitude'],
                    'check_out_lng': log['longitude'],
                    'duration': None
                })
                
    all_shifts = []
    for uid, s_list in shifts_by_user.items():
        all_shifts.extend(s_list)
        
    def get_sort_key(s):
        return s['check_in'] or s['check_out'] or "1970-01-01 00:00:00"
        
    all_shifts.sort(key=get_sort_key, reverse=True)
    recent_shifts = all_shifts[:10]
    
    conn.close()
    
    return jsonify({
        'active_staff_count': active_staff_count,
        'today_checklist_count': today_checklist_count,
        'total_areas': total_areas,
        'total_staff': total_staff,
        'recent_logs': recent_shifts
    })

# Quản lý nhân viên
@app.route('/api/manager/staff', methods=['GET', 'POST'])
def manager_staff():
    conn = get_db()
    if request.method == 'GET':
        staff = conn.execute('''
            SELECT u.*, a.name as area_name 
            FROM users u 
            LEFT JOIN areas a ON u.area_id = a.id 
            ORDER BY u.role, u.name
        ''').fetchall()
        conn.close()
        return jsonify([dict(s) for s in staff])
        
    elif request.method == 'POST':
        data = request.json or {}
        uid = data.get('id')
        name = data.get('name', '').strip()
        code = data.get('code', '').strip().upper()
        role = data.get('role', 'fulltime')
        cccd = data.get('cccd', '').strip()
        area_id = data.get('area_id')
        
        if not name or not code:
            conn.close()
            return jsonify({'error': 'Họ tên và Mã đăng nhập không được trống'}), 400
            
        # Bắt buộc số CCCD đối với Fulltime và Parttime
        if role in ['fulltime', 'parttime']:
            if not cccd:
                conn.close()
                return jsonify({'error': 'Số CCCD là thông tin bắt buộc đối với nhân sự Fulltime và Parttime.'}), 400
            import re
            if not re.match(r'^\d{12}$', cccd):
                conn.close()
                return jsonify({'error': 'Số CCCD không hợp lệ. Vui lòng nhập đúng 12 chữ số.'}), 400
                
            # Kiểm tra CCCD trùng lặp
            if uid:
                existing_cccd = conn.execute("SELECT id FROM users WHERE cccd = ? AND id != ?", (cccd, uid)).fetchone()
            else:
                existing_cccd = conn.execute("SELECT id FROM users WHERE cccd = ?", (cccd,)).fetchone()
            if existing_cccd:
                conn.close()
                return jsonify({'error': 'Số CCCD này đã tồn tại trên hệ thống. Vui lòng kiểm tra lại.'}), 400
        
        try:
            if uid: # Cập nhật
                conn.execute('''
                    UPDATE users 
                    SET name = ?, code = ?, role = ?, cccd = ?, area_id = ?
                    WHERE id = ?
                ''', (name, code, role, cccd if role != 'manager' else (cccd or None), area_id if role != 'manager' else None, uid))
            else: # Thêm mới
                conn.execute('''
                    INSERT INTO users (name, code, role, cccd, area_id)
                    VALUES (?, ?, ?, ?, ?)
                ''', (name, code, role, cccd if role != 'manager' else (cccd or None), area_id if role != 'manager' else None))
            conn.commit()
            conn.close()
            return jsonify({'success': True})
        except sqlite3.IntegrityError:
            conn.close()
            return jsonify({'error': 'Mã đăng nhập đã tồn tại trên hệ thống'}), 400

@app.route('/api/manager/staff/delete', methods=['POST'])
def manager_staff_delete():
    data = request.json or {}
    uid = data.get('id')
    if not uid:
        return jsonify({'error': 'Thiếu ID nhân sự'}), 400
    conn = get_db()
    conn.execute("UPDATE users SET status = 'inactive' WHERE id = ?", (uid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# Quản lý khu vực
@app.route('/api/manager/areas', methods=['GET', 'POST'])
def manager_areas():
    conn = get_db()
    if request.method == 'GET':
        areas = conn.execute("SELECT * FROM areas ORDER BY name").fetchall()
        conn.close()
        return jsonify([dict(a) for a in areas])
        
    elif request.method == 'POST':
        data = request.json or {}
        aid = data.get('id')
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        
        if not name:
            conn.close()
            return jsonify({'error': 'Tên khu vực không được trống'}), 400
            
        try:
            if aid:
                conn.execute("UPDATE areas SET name = ?, description = ? WHERE id = ?", (name, description, aid))
            else:
                conn.execute("INSERT INTO areas (name, description) VALUES (?, ?)", (name, description))
            conn.commit()
            conn.close()
            return jsonify({'success': True})
        except sqlite3.IntegrityError:
            conn.close()
            return jsonify({'error': 'Tên khu vực đã tồn tại'}), 400

@app.route('/api/manager/areas/delete', methods=['POST'])
def manager_areas_delete():
    data = request.json or {}
    aid = data.get('id')
    if not aid:
        return jsonify({'error': 'Thiếu ID khu vực'}), 400
    conn = get_db()
    conn.execute("DELETE FROM areas WHERE id = ?", (aid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# Quản lý quy tắc chấm chéo
@app.route('/api/manager/cross_check_rules', methods=['GET', 'POST'])
def manager_cross_check_rules():
    conn = get_db()
    if request.method == 'GET':
        rules = conn.execute('''
            SELECT r.*, a1.name as from_area_name, a2.name as to_area_name 
            FROM cross_check_rules r
            JOIN areas a1 ON r.from_area_id = a1.id
            JOIN areas a2 ON r.to_area_id = a2.id
            ORDER BY a1.name
        ''').fetchall()
        conn.close()
        return jsonify([dict(r) for r in rules])
        
    elif request.method == 'POST':
        data = request.json or {}
        from_area_id = data.get('from_area_id')
        to_area_id = data.get('to_area_id')
        
        if not from_area_id or not to_area_id:
            conn.close()
            return jsonify({'error': 'Vui lòng chọn đầy đủ bộ phận đi chấm và bộ phận được chấm'}), 400
            
        if int(from_area_id) == int(to_area_id):
            conn.close()
            return jsonify({'error': 'Một bộ phận không thể tự chấm chéo chính mình. Vui lòng chọn bộ phận khác.'}), 400
            
        try:
            conn.execute('''
                INSERT OR REPLACE INTO cross_check_rules (from_area_id, to_area_id)
                VALUES (?, ?)
            ''', (from_area_id, to_area_id))
            conn.commit()
            conn.close()
            return jsonify({'success': True})
        except Exception as e:
            conn.close()
            return jsonify({'error': f'Lỗi khi lưu quy tắc: {str(e)}'}), 500

@app.route('/api/manager/cross_check_rules/delete', methods=['POST'])
def manager_cross_check_rules_delete():
    data = request.json or {}
    from_area_id = data.get('from_area_id')
    if not from_area_id:
        return jsonify({'error': 'Thiếu ID bộ phận đi chấm'}), 400
    conn = get_db()
    conn.execute("DELETE FROM cross_check_rules WHERE from_area_id = ?", (from_area_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# Quản lý hạng mục checklist
@app.route('/api/manager/checklist_items', methods=['GET', 'POST'])
def manager_checklist_items():
    conn = get_db()
    if request.method == 'GET':
        area_id = request.args.get('area_id', type=int)
        if area_id:
            items = conn.execute("SELECT * FROM checklist_items WHERE area_id = ?", (area_id,)).fetchall()
        else:
            items = conn.execute('''
                SELECT i.*, a.name as area_name 
                FROM checklist_items i
                JOIN areas a ON i.area_id = a.id
                ORDER BY a.name, i.task_name
            ''').fetchall()
        conn.close()
        return jsonify([dict(item) for item in items])
        
    elif request.method == 'POST':
        # Vì có upload ảnh nên nhận dạng form-data
        item_id = request.form.get('id')
        area_id = request.form.get('area_id')
        task_name = request.form.get('task_name', '').strip()
        
        if not area_id or not task_name:
            conn.close()
            return jsonify({'error': 'Vui lòng điền đầy đủ thông tin'}), 400
            
        # Xử lý file ảnh mẫu
        ref_image_url = None
        if 'image' in request.files:
            file = request.files['image']
            if file.filename != '':
                filename = secure_filename(file.filename)
                ext = os.path.splitext(filename)[1]
                filename = f"ref_{int(datetime.datetime.now().timestamp())}{ext}"
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(file_path)
                ref_image_url = f"/static/uploads/{filename}"
                
        if item_id: # Update
            if ref_image_url:
                conn.execute('''
                    UPDATE checklist_items 
                    SET area_id = ?, task_name = ?, reference_image = ?
                    WHERE id = ?
                ''', (area_id, task_name, ref_image_url, item_id))
            else:
                conn.execute('''
                    UPDATE checklist_items 
                    SET area_id = ?, task_name = ?
                    WHERE id = ?
                ''', (area_id, task_name, item_id))
        else: # Add new
            conn.execute('''
                INSERT INTO checklist_items (area_id, task_name, reference_image)
                VALUES (?, ?, ?)
            ''', (area_id, task_name, ref_image_url))
            
        conn.commit()
        conn.close()
        return jsonify({'success': True})

@app.route('/api/manager/checklist_items/delete', methods=['POST'])
def manager_checklist_items_delete():
    data = request.json or {}
    item_id = data.get('id')
    if not item_id:
        return jsonify({'error': 'Thiếu ID hạng mục'}), 400
    conn = get_db()
    conn.execute("DELETE FROM checklist_items WHERE id = ?", (item_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# Xem danh sách checklist đã nộp để phê duyệt
@app.route('/api/manager/submissions', methods=['GET'])
def manager_submissions():
    status = request.args.get('status', 'pending')
    conn = get_db()
    
    subs = conn.execute('''
        SELECT s.*, u.name as grader_name, a.name as area_name
        FROM checklist_submissions s
        JOIN users u ON s.grader_id = u.id
        JOIN areas a ON s.area_id = a.id
        WHERE s.status = ?
        ORDER BY s.timestamp DESC
    ''', (status,)).fetchall()
    
    res = []
    for sub in subs:
        details = conn.execute('''
            SELECT i.task_name, i.reference_image, d.status, d.captured_image, d.notes, d.id as detail_id
            FROM checklist_items i
            LEFT JOIN checklist_submission_details d ON i.id = d.item_id AND d.submission_id = ?
            WHERE i.area_id = ?
        ''', (sub['id'], sub['area_id'])).fetchall()
        
        res.append({
            'id': sub['id'],
            'grader_name': sub['grader_name'],
            'area_name': sub['area_name'],
            'timestamp': sub['timestamp'],
            'status': sub['status'],
            'manager_notes': sub['manager_notes'],
            'latitude': sub['latitude'],
            'longitude': sub['longitude'],
            'details': [dict(d) for d in details]
        })
        
    conn.close()
    return jsonify(res)

@app.route('/api/manager/submissions/approve', methods=['POST'])
def manager_submissions_approve():
    data = request.json or {}
    sub_id = data.get('submission_id')
    status = data.get('status') # 'approved' hoặc 'rejected'
    notes = data.get('notes', '').strip()
    
    if not sub_id or not status:
        return jsonify({'error': 'Thiếu dữ liệu phê duyệt'}), 400
        
    conn = get_db()
    conn.execute('''
        UPDATE checklist_submissions 
        SET status = ?, manager_notes = ?
        WHERE id = ?
    ''', (status, notes, sub_id))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

# Báo cáo hiệu suất
@app.route('/api/manager/reports/efficiency', methods=['GET'])
def api_reports_efficiency():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search_query = request.args.get('search')
    data = get_staff_kpi_data(start_date, end_date, search_query)
    return jsonify(data)

# Xuất Excel: Chấm công
@app.route('/api/manager/reports/export/attendance', methods=['GET'])
def export_attendance():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search_query = request.args.get('search')
    
    conn = get_db()
    query_params = []
    
    sql = '''
        SELECT l.*, u.name as user_name, u.code as user_code, u.role as user_role, a.name as area_name
        FROM time_logs l
        JOIN users u ON l.user_id = u.id
        LEFT JOIN areas a ON l.area_id = a.id
        WHERE u.role != 'manager'
    '''
    if start_date and end_date:
        sql += " AND date(l.timestamp) >= ? AND date(l.timestamp) <= ?"
        query_params.extend([start_date, end_date])
        
    if search_query:
        sql += " AND (u.name LIKE ? OR u.code LIKE ?)"
        q = f"%{search_query}%"
        query_params.extend([q, q])
        
    sql += " ORDER BY l.timestamp DESC"
    logs = conn.execute(sql, query_params).fetchall()
    conn.close()
    
    # Khởi tạo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Bao Cao Cham Cong"
    
    # Style
    font_title = Font(name="Calibri", size=16, bold=True, color="1B5E20")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    fill_header = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    border_thin = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Viết tiêu đề
    ws.merge_cells("A1:G2")
    title_cell = ws["A1"]
    title_cell.value = "BÁO CÁO CHI TIẾT LỊCH SỬ CHẤM CÔNG NHÂN VIÊN"
    title_cell.font = font_title
    title_cell.alignment = align_center
    
    # Headers
    headers = ["STT", "Mã Nhân Viên", "Tên Nhân Viên", "Chức Vụ", "Hành Động", "Thời Gian", "Khu Vực Làm Việc"]
    ws.append([]) # Dòng 3 trống
    ws.append(headers) # Dòng 4
    
    for col_num in range(1, 8):
        cell = ws.cell(row=4, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin
        
    # Thêm dữ liệu
    row_idx = 5
    for idx, log in enumerate(logs):
        role_map = {'manager': 'Quản lý', 'fulltime': 'Fulltime', 'parttime': 'Parttime'}
        action_map = {'check_in': 'Vào ca (Check-in)', 'check_out': 'Ra ca (Check-out)'}
        
        row_data = [
            idx + 1,
            log['user_code'],
            log['user_name'],
            role_map.get(log['user_role'], log['user_role']),
            action_map.get(log['action'], log['action']),
            log['timestamp'],
            log['area_name'] or "-"
        ]
        ws.append(row_data)
        
        for col_num in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.border = border_thin
            if col_num in [1, 2, 4, 5, 6]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
        row_idx += 1
        
    # Thiết lập độ rộng cột tự động
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Lưu file gửi về client
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Bao_Cao_Cham_Cong.xlsx"
    )

# Xuất Excel: Hiệu suất & KPIs
@app.route('/api/manager/reports/export/efficiency', methods=['GET'])
def export_efficiency():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search_query = request.args.get('search')
    kpi_data = get_staff_kpi_data(start_date, end_date, search_query)
    
    # Khởi tạo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Hieu Suat Ca & KPIs"
    
    # Style
    font_title = Font(name="Calibri", size=16, bold=True, color="0D47A1")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    fill_header = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    border_thin = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Tiêu đề
    ws.merge_cells("A1:I2")
    title_cell = ws["A1"]
    title_cell.value = "BẢO CÁO ĐÁNH GIÁ HIỆU SUẤT & KPIs NHÂN SỰ"
    title_cell.font = font_title
    title_cell.alignment = align_center
    
    # Headers
    headers = [
        "STT", "Mã Nhân Viên", "Tên Nhân Viên", "Khu Vực Phụ Trách", 
        "Tổng Giờ Làm", "Số Lần Check-in", "Số Lượt Chấm Checklist", 
        "Điểm Vệ Sinh (%)", "Tỷ Lệ Chấm Chéo (%)", "Đánh Giá Hiệu Suất"
    ]
    ws.append([]) # Dòng 3
    ws.append(headers) # Dòng 4
    
    for col_num in range(1, 11):
        cell = ws.cell(row=4, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin
        
    # Thêm dữ liệu
    row_idx = 5
    for idx, item in enumerate(kpi_data):
        row_data = [
            idx + 1,
            item['code'],
            item['name'],
            item['area_name'],
            item['total_hours'],
            item['checkins_count'],
            item['checklists_submitted'],
            f"{item['hygiene_score']}%",
            f"{item['completion_rate']}%",
            item['status']
        ]
        ws.append(row_data)
        
        # Tô màu trạng thái
        status_cell = ws.cell(row=row_idx, column=10)
        if item['status'] == "Hiệu quả":
            status_cell.font = Font(name="Calibri", size=11, bold=True, color="2E7D32")
        else:
            status_cell.font = Font(name="Calibri", size=11, bold=True, color="C62828")
            
        for col_num in range(1, 11):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.border = border_thin
            if col_num in [1, 2, 5, 6, 7, 8, 9, 10]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
        row_idx += 1
        
    # Thiết lập độ rộng cột tự động
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Lưu file gửi về client
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Bao_Cao_Hieu_Suat_Nhan_Su.xlsx"
    )

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
